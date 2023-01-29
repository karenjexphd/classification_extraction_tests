# Extract csv from xlsx - one csv file per sheet

from openpyxl import load_workbook
import csv
import sys

infilepath = str(sys.argv[1])
filename = str(sys.argv[2])
outfilepath = str(sys.argv[3])

#basename= filename.split('.')[0]

full_filename=infilepath+'/'+filename

wb = load_workbook(full_filename)

# print ('sheet names:',wb.sheetnames)

for sheet in wb.sheetnames:
    # csv_filename=filepath+'/'+basename+'_'+sheet+'.csv'
    csv_filename=outfilepath+'/'+sheet+'.csv'
    # print('ws:',ws)
    ws=wb[sheet]
    data=list(ws.rows)

    col = csv.writer(open(csv_filename,
                          'w',
                          newline=""))
    for r in ws.rows:
        # row by row write 
        col.writerow([cell.value for cell in r])
