from openpyxl import load_workbook
import sys 
import os
import psycopg2
import string

# Goals: 
#   For each table in each input file, insert a row into source_table
#   The row will contain the coordinates of the first and last cell in the table
#   Based on the $START and $END annotated cells
#   Note that $START is above and to the left of the first cell in the table
#             $END is below and to the right of the last cell in the table
#   if $START is in cell A2, i.e. the cell with coordinates (1,2), table_first_col = 2 and table_first_row=3
#   if $END is in cell F10, i.e. the cell with coordinates (6,10), table_last_col = 5 and table_last_row=9

# 1. Process input parameters:

#   i. input_filepath       path to file to be processed
input_filepath = str(sys.argv[1])               
# input_filepath = '/home/karen/workspaces/classification_extraction_tests/test_files/tabby_10_files/xlsx'

# sheet_num and table_num will always be 0 for now - assuming one sheet per file, and one table per sheet
sheet_num=0
table_num=0

# 2. Create connection to table_model database with search_path set to table_model
#    (Need to parameterise this)

tm_conn = psycopg2.connect(
    host="127.0.0.1",
    database="table_model",
    user="postgres")

cur = tm_conn.cursor()
cur.execute('SET SEARCH_PATH=table_model')

def address2coords(cell_address):
    # Function to return (numeric) coordinates of a spreadsheet cell based on its address
    # A3 will return (1,3)  AA43 will return (27,43)
    col=''.join(filter(str.isalpha, cell_address))   # alpha part of cell address
    row=int(''.join(filter(str.isdigit, cell_address)))  # numeric part of cell address
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num,row

for file in os.listdir(input_filepath):
    input_file = input_filepath+"/"+file      # fully qualified file
    base_filename=file.split('.xlsx')[0]      # base filename

    # Use openpyxl load_workbook to load the input file as a workbook (wb)
    # and get data from the required sheet as a worksheet (ws):

    wb = load_workbook(input_file)
    sheets=wb.sheetnames
    ws=wb[sheets[sheet_num]]
    
    all_rows = list(ws.rows)
    for row in all_rows:
        for cell in row:
            if cell.value == "$START":
                table_start=cell.coordinate
            elif cell.value == "$END":
                table_end=cell.coordinate

    table_first_col=address2coords(table_start)[0]+1     # first col in table is to right of table_start
    table_first_row=address2coords(table_start)[1]+1     # first row in table is below table_start

    table_last_col=address2coords(table_end)[0]-1     # last col in table is to left of table_end
    table_last_row=address2coords(table_end)[1]-1     # last row in table is above of table_end

    # tablestart_col=''.join(filter(str.isalpha, table_start))        # alpha part, e.g. "A" from "A2"
    # tablestart_row=int(''.join(filter(str.isdigit, table_start)))   # numeric part, e.g. 2 from "A2"
    # tableend_col=''.join(filter(str.isalpha, table_end))        # alpha part, e.g. "U" from "U2"
    # tableend_row=int(''.join(filter(str.isdigit, table_end)))   # numeric part, e.g. 12 from "U12"

    # Populate source_table
    # with a single row describing the table being processed

    insert_stmt="INSERT INTO source_table ( \
                table_is_gt, \
                table_method, \
                table_first_col, \
                table_first_row, \
                table_last_col, \
                table_last_row, \
                file_name,  \
                sheet_number, \
                table_number) \
                VALUES (TRUE, \
                        'tabbyxl', \
                        "+str(table_first_col)+", \
                        "+str(table_first_row)+", \
                        "+str(table_last_col)+", \
                        "+str(table_last_row)+", \
                        '"+base_filename+"', \
                        "+str(sheet_num)+", \
                        "+str(table_num)+")"
    cur.execute(insert_stmt)

# Commit changes to retain data input into tables
cur.execute('COMMIT;')
cur.close()