from openpyxl import load_workbook
import sys 
import psycopg2
import string

# Goals: 
#   Extract information from given Tabby ground truth file or output file and map it to table model
#   Note: a tabby GT or output file contains a single table

# 1. Process input parameters:

#   i. input_filepath       path to file to be processed
input_filepath = str(sys.argv[1])               
#   ii. input_filename       name of file to be processed (<filename>[_sheetnum]_[tablenum].xlsx)
input_filename = str(sys.argv[2])               
#   iii. is_gt                TRUE if this is a file containing ground truth, FALSE if is is an output file 
if str(sys.argv[3]) == 'TRUE':
    is_gt=True
else:
    is_gt=False

input_file = input_filepath+"/"+input_filename  # Fully qualified file

#sample input_file for testing
#input_file="/home/karen/workspaces/classification_extraction_tests/test_files/tabby_small_file/gt/tabby/smpl_0_0.xlsx"
#input_file="/home/karen/workspaces/classification_extraction_tests/test_files/tabby_10_files/gt/C10001.xlsx"

# Get base filename based on input_file path and name
base_filename=input_filename.split('.xlsx')[0]

filename= base_filename
sheetnum=0
tablenum=0    

# 2. Create connection to table_model database with search_path set to table_model
#    (Need to parameterise this)

tm_conn = psycopg2.connect(
    host="127.0.0.1",
    database="table_model",
    user="postgres")

cur = tm_conn.cursor()
cur.execute('SET SEARCH_PATH=table_model')


def address2coords(cell_address):
    # Function to return (numeric) coordinates of a spreadsheet cell based on its address
    # A3 will return (1,3)  AA43 will return (27,43)
    col=''.join(filter(str.isalpha, cell_address))   # alpha part of cell address
    row=int(''.join(filter(str.isdigit, cell_address)))  # numeric part of cell address
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num,row

# retrieve existing data (table ID, table first/last col & row) from source_table for the file/sheet/table being processed

select_stmt="SELECT table_id, \
                    table_first_col, \
                    table_first_row, \
                    table_last_col, \
                    table_last_row \
            FROM source_table \
            WHERE table_is_gt=TRUE \
            AND file_name='"+filename+"' \
            AND sheet_number="+str(sheetnum)+" \
            AND table_number="+str(tablenum)
cur.execute(select_stmt)

table_info = cur.fetchone()

gt_table_id    = table_info[0]
table_first_col = table_info[1]
table_first_row = table_info[2]
table_last_col   = table_info[3]
table_last_row   = table_info[4]

if is_gt:

    # We're processing the ground truth file so source_table is already populated
    # Set gt_table_id to table_id in preparation for populating the remaining tables
    table_id=gt_table_id

else:       
    
    # We're processing the output file so we need to populate source_table with the information just retrieved

    insert_stmt="INSERT INTO source_table ( \
                    table_is_gt, \
                    table_method, \
                    table_first_col, \
                    table_first_row, \
                    table_last_col, \
                    table_last_row, \
                    file_name, \
                    sheet_number, \
                    table_number) \
                VALUES ( \
                    FALSE, \
                    'tabbyxl', \
                    "+str(table_first_col)+", \
                    "+str(table_first_row)+", \
                    "+str(table_last_col)+", \
                    "+str(table_last_row)+", \
                    '"+filename+"', \
                    "+str(sheetnum)+", \
                    "+str(tablenum)+")"
    cur.execute(insert_stmt)

    # retrieve (automatically generated) table_id from source_table

    select_stmt="SELECT table_id FROM source_table \
                WHERE table_is_gt=FALSE \
                AND table_method='tabbyxl' \
                AND file_name='"+filename+"' \
                AND sheet_number="+str(sheetnum)+" \
                AND table_number="+str(tablenum)
    cur.execute(select_stmt)
    
    table_info = cur.fetchone() 

    table_id = table_info[0]
    print('table_id: '+str(table_id))

# Continue to populate the table_model whether this is a ground truth or an output file

# Use openpyxl's load_workbook to load the file and get data from the ENTRIES and LABELS sheets:

wb         = load_workbook(input_file)

ws_entries = wb['ENTRIES']
ws_labels  = wb['LABELS']
 
all_entries_rows = list(ws_entries.rows)
all_labels_rows = list(ws_labels.rows)

# Populate temporary tables entry_temp and entry_label_temp from rows in ENTRIES

# NOTE: header row not processed - currently checking each row to see if it's the header - improve this!

for row in all_entries_rows:
    # get value from ENTRY column (row[0].value) as entry_val
    entry_val  = str(row[0].value)
    entry_prov_text = str(row[1].value)                       # value from PROVENANCE column
    entry_labels = str(row[2].value)                          # value from LABELS column
    if entry_prov_text != 'PROVENANCE':                       # ignore header row
        # The provenance column contains a hyperlink and a display val

        # Extract just the display val, i.e. the part between "," and ")
        entry_prov=entry_prov_text.split('","')[1].split('")')[0]   

        # The display val is a cell reference, e.g. C3. Convert to cell and row IDs
        entry_prov_col=address2coords(entry_prov)[0]
        entry_prov_row=address2coords(entry_prov)[1]

        # entry_prov_col=''.join(filter(str.isalpha, entry_prov))        # alpha part
        # entry_prov_row=int(''.join(filter(str.isdigit, entry_prov)))   # numeric part

        insert_et="INSERT INTO entry_temp (\
                    table_id, \
                    entry_value, \
                    entry_provenance, \
                    entry_provenance_col, \
                    entry_provenance_row) \
                   VALUES ( \
                    "+str(table_id)+", \
                    '"+str(entry_val)+"', \
                    '"+entry_prov+"', \
                    "+str(entry_prov_col)+", \
                    "+str(entry_prov_row)+")"
        cur.execute(insert_et)
        # split entry_labels (on comma surrounded by double quotes) to get list of labels for this entry
        entry_label_list = entry_labels.split('", "')
        for entry_label in entry_label_list:
            # label_provenance is the cell reference found between square brackets IF THE LABEL IS NOT AN EMPTY STRING
            if entry_label == '':
                label_prov=''
            else:
              label_prov=entry_label.split('[')[1].split(']')[0]

            # insert record into temp table entry_label_temp for this label
            insert_stmt_elt="INSERT INTO entry_label_temp (table_id, entry_provenance, label_provenance) \
                             VALUES ("+str(table_id)+",'"+entry_prov+"', '"+label_prov+"')"
            cur.execute(insert_stmt_elt)

# Populate table_cell based on contents of entry_temp

# The table_cell coordinates (left_col and top_row) represent the position of the cell within the table
# The values are calculated from entry_provenance (the cell's position within the file) 
# and the position of the first cell in the table within the file

#   NOTES: 
#     i.  The GT doesn't contain right_col, bottom_row or cell_datatype
#         right_col and bottom_row are therefore inserted as duplicates of left_col and top_row
#         cell_datatype is left empty

insert_stmt="INSERT INTO table_cell (\
                    table_id, \
                    left_col, \
                    top_row, \
                    right_col, \
                    bottom_row, \
                    cell_content, \
                    cell_annotation) \
            SELECT  "+str(table_id)+", \
                    entry_provenance_col-"+str(table_first_col)+", \
                    entry_provenance_row-"+str(table_first_row)+", \
                    entry_provenance_col-"+str(table_first_col)+", \
                    entry_provenance_row-"+str(table_first_row)+", \
                    entry_value, 'DATA' \
            FROM entry_temp WHERE table_id="+str(table_id)
cur.execute(insert_stmt)

# Populate entry based on table_cell

insert_stmt="INSERT INTO entry (entry_cell_id) \
             SELECT cell_id FROM table_cell \
             WHERE table_id="+str(table_id)+" AND cell_annotation='DATA'"
cur.execute(insert_stmt)

# Process each row from the LABELS sheet (all_labels_rows)
# and insert the data into the temporary table label_temp
# NOTES: 
#    header row not processed - currently checking each row to see if it's the header - improve this!
#    labels containing single quotes cause errors. Either surround text in double quotes or escape the single quotes

for row in all_labels_rows:
    label_val  = str(row[0].value).replace("'","''")           # value in LABEL column with single quotes escaped
    label_val = label_val
    label_prov_text = str(row[1].value)                        # value in PROVENANCE column
    label_par  = str(row[2].value)                             # value in PARENT column
    label_cat  = str(row[3].value)                             # value in CATEGORY column
    if label_prov_text != 'PROVENANCE':                        # ignore header row
        label_prov=label_prov_text.split('","')[1].split('")')[0]      # get display val between "," and ")

        label_prov_col=address2coords(label_prov)[0]
        label_prov_row=address2coords(label_prov)[1]

        # label_prov_col=''.join(filter(str.isalpha, label_prov))        # alpha part
        # label_prov_row=int(''.join(filter(str.isdigit, label_prov)))   # numeric part        

        # If the label has no parent labels, insert a row containing information for just this label
        if label_par=='None':
            insert_stmt="INSERT INTO label_temp (table_id, label_value, label_provenance, label_provenance_col, label_provenance_row, label_category) \
                        VALUES ("+str(table_id)+",'"+label_val+"', '"+label_prov+"', '"+str(label_prov_col)+"', '"+str(label_prov_row)+"', '"+label_cat+"')"
        # If the label has a parent, insert a row that includes provenance information for the parent label
        else:
            label_par_prov=label_par.split('[')[1].split(']')[0]           # get val between '[' and ']'
            insert_stmt="INSERT INTO label_temp (table_id, label_value, label_provenance, label_provenance_col, label_provenance_row, label_parent, label_category) \
                        VALUES ("+str(table_id)+",'"+label_val+"', '"+label_prov+"', '"+str(label_prov_col)+"', '"+str(label_prov_row)+"', '"+label_par_prov+"', '"+label_cat+"')"
        cur.execute(insert_stmt)

# Populate the category table from the label_categories in label_temp

insert_stmt_cat="INSERT INTO category (category_name, table_id) \
                 SELECT DISTINCT label_category, "+str(table_id)+" FROM label_temp WHERE table_id="+str(table_id)
cur.execute(insert_stmt_cat)

# Populate table_cell based on label_temp contents
# The table_cell coordinates are calculated as for the table_cell rows corresponding to entries

insert_stmt="INSERT INTO table_cell \
            (table_id, left_col, top_row, right_col, bottom_row, cell_content, cell_annotation) \
            SELECT  "+str(table_id)+", \
                    label_provenance_col-"+str(table_first_col)+", \
                    label_provenance_row-"+str(table_first_row)+", \
                    label_provenance_col-"+str(table_first_col)+", \
                    label_provenance_row-"+str(table_first_row)+", \
                    label_value, 'HEADING' \
            FROM label_temp WHERE table_id="+str(table_id)
cur.execute(insert_stmt)

# Populate label based on table_cell and label_temp

insert_stmt="INSERT INTO label (label_cell_id, category_name, parent_label_cell_id) \
             SELECT tcv.cell_id, lt.label_category, tpcv.cell_id \
             FROM label_temp lt \
             JOIN tabby_cell_view tcv \
             ON lt.label_provenance=tcv.cell_provenance  \
             AND lt.table_id=tcv.table_id \
             LEFT JOIN tabby_cell_view tpcv \
             ON lt.label_parent=tpcv.cell_provenance  \
             AND lt.table_id = tpcv.table_id \
             WHERE lt.table_id="+str(table_id)+" AND tcv.cell_annotation='HEADING'"
cur.execute(insert_stmt)

# Populate entry_label from entry_label_temp and tabby_cell_view

# NEED TO ADD IN table_ID?

insert_el="INSERT INTO entry_label (entry_cell_id, label_cell_id) \
           SELECT e_cell.cell_id, l_cell.cell_id \
           FROM entry_label_temp elt \
           JOIN tabby_cell_view e_cell ON elt.entry_provenance = e_cell.cell_provenance \
           JOIN tabby_cell_view l_cell ON elt.label_provenance = l_cell.cell_provenance \
           WHERE elt.table_id="+str(table_id)+" \
           AND e_cell.table_id="+str(table_id)+" \
           AND l_cell.table_id="+str(table_id)
cur.execute(insert_el)

# Empty temp tables

truncate_et="TRUNCATE TABLE entry_temp, label_temp, entry_label_temp"
cur.execute(truncate_et)

# Commit changes to retain data input into tables
cur.execute('COMMIT;')
cur.close()