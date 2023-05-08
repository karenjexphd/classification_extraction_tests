from openpyxl import load_workbook
import sys 
import psycopg2

# Goals: 
#   Extract information from given Tabby ground truth file and map it to table model
#   Display full & correct information for the processed table in the view canonical_table_view
#   This can then be compared to the canonical_table_view extracted during the table extraction process

# 1. Process input (Tabby ground truth) file (filename_<sheetnum>_<tablenum>.xlsx)

#    Note: a tabby GT file contains a single table

#    input_file has been hardcoded during unit testing. 
#    Will be provided as input parameter during end-to-end tests:

#sample input_file for testing
#input_file="/home/karen/workspaces/classification_extraction_tests/test_files/tabby_small_file/gt/tabby/smpl_0_0.xlsx"
#input_file="/home/karen/workspaces/classification_extraction_tests/test_files/tabby_10_files/gt/C10001.xlsx"

input_filepath = str(sys.argv[1])               # path to TabbyXL format GT file
input_filename = str(sys.argv[2])               # Name of TabbyXL format GT file (<filename>[_sheetnum]_[tablenum].xlsx)
input_file = input_filepath+"/"+input_filename  # Fully qualified TabbyXL format GT file

# Get base filename based on input_file path and name
base_filename=input_filename.split('.xlsx')[0]

#    Identify filename, sheet number and table number
#    If len(base_filename.split('_')==3 ) then base_filename=filename_sheetnum_tablenum
#    Else, sheetnum and tablenum are both 0

if len(base_filename.split('_'))==3:
    filename=base_filename.split('_')[0]
    sheetnum=base_filename.split('_')[1]
    tablenum=base_filename.split('_')[2]
else:
    filename=base_filename
    sheetnum=0
    tablenum=0    
    
#    table_start and table_end have been hardcoded during unit testing. 
#    Will be provided as input parameter during end-to-end tests
#    (will have been identified in a previous step from the original input_file):

tablestart='A2'
tableend='U12'

tablestart_col=''.join(filter(str.isalpha, tablestart))        # alpha part, e.g. "A" from "A2"
tablestart_row=int(''.join(filter(str.isdigit, tablestart)))   # numeric part, e.g. 12 from "U12"

tableend_col=''.join(filter(str.isalpha, tableend))        # alpha part, e.g. "A" from "A2"
tableend_row=int(''.join(filter(str.isdigit, tableend)))   # numeric part, e.g. 12 from "U12"

# 2. Create connection to table_model database with search_path set to table_model
#    (Need to parameterise this)

tm_conn = psycopg2.connect(
    host="p.qnplnpl3nbabto2zddq2phjlwi.db.postgresbridge.com",
    database="table_model",
    user="postgres")

cur = tm_conn.cursor()
cur.execute('SET SEARCH_PATH=table_model')

# Populate source_table
# with a single row describing the table being processed

insert_stmt="INSERT INTO source_table (table_start_col, table_start_row, table_end_col, table_end_row, file_name, sheet_number, table_number) \
             VALUES ('"+tablestart_col+"', "+str(tablestart_row)+", '"+tableend_col+"', "+str(tableend_row)+", '"+filename+"', "+str(sheetnum)+", "+str(tablenum)+")"
cur.execute(insert_stmt)

# retrieve (automatically generated) table_id from source_table

select_stmt="SELECT table_id FROM source_table \
             WHERE file_name='"+filename+"' AND sheet_number="+str(sheetnum)+" AND table_number="+str(tablenum)
cur.execute(select_stmt)
table_id = cur.fetchone()[0]
print('table_id: '+str(table_id))

# Use openpyxl load_workbook to load the input file as a workbook 
# and get data from the required sheets (ENTRIES and LABELS):

wb         = load_workbook(input_file)

ws_entries = wb['ENTRIES']
ws_labels  = wb['LABELS']
 
all_entries_rows = list(ws_entries.rows)
all_labels_rows = list(ws_labels.rows)

# Populate temporary tables entry_temp and entry_label_temp
# from the rows in the ENTRIES sheet

# Process each row from the ENTRIES sheet (all_entries_rows)
# and insert the data into the temporary table entry_temp

# NOTE: header row not processed - currently checking each row to see if it's the header - improve this!

for row in all_entries_rows:
    entry_val  = str(row[0].value)                            # value from ENTRY column
    entry_prov_text = str(row[1].value)                       # value from PROVENANCE column
    entry_labels = str(row[2].value)                          # value from LABELS column
    if entry_prov_text != 'PROVENANCE':                       # ignore header row
        # The provenance column contains a hyperlink and a display val
        # Extract just the display val, i.e. the part between "," and ")
        entry_prov=entry_prov_text.split('","')[1].split('")')[0]   
        # The display val is a cell reference, e.g. C3. Split this into its alpha and numeric parts
        entry_prov_col=''.join(filter(str.isalpha, entry_prov))        # alpha part
        entry_prov_row=int(''.join(filter(str.isdigit, entry_prov)))   # numeric part

        # Have removed entry_labels from the following insert statement as the info is not used elsewhere

        # insert_et="INSERT INTO entry_temp (entry_value, entry_provenance, entry_provenance_col, entry_provenance_row, entry_labels) \
        #              VALUES ('"+entry_val+"', '"+entry_prov+"', '"+entry_prov_col+"', "+str(entry_prov_row)+", '"+entry_labels+"')"

        insert_et="INSERT INTO entry_temp (table_id, entry_value, entry_provenance, entry_provenance_col, entry_provenance_row) \
                     VALUES ("+str(table_id)+",'"+entry_val+"', '"+entry_prov+"', '"+entry_prov_col+"', "+str(entry_prov_row)+")"
        cur.execute(insert_et)
        # split entry_labels (on comma) to get list of labels for this entry
        entry_label_list = entry_labels.split('", "')
        for entry_label in entry_label_list:
            # label_provenance is the cell reference found between square brackets
            label_prov=entry_label.split('[')[1].split(']')[0]
            #print('table '+str(table_id)+' entry_label'+entry_label+' entry_prov '+entry_prov+' label_prov '+label_prov)
            # insert record into temp table entry_label_temp for this label
            insert_stmt_elt="INSERT INTO entry_label_temp (table_id, entry_provenance, label_provenance) \
                             VALUES ("+str(table_id)+",'"+entry_prov+"', '"+label_prov+"')"
            cur.execute(insert_stmt_elt)

# Populate table_cell based on contents of entry_temp

# The table_cell coordinates (left_col and top_row) represent the position of the cell within the table
# The values are calculated from entry_provenance (the cell's position within the file) 
# and tablestart, i.e. the position of the start of the table within the file

#   NOTES: 
#     i.  The GT doesn't contain right_col, bottom_row or cell_datatype
#         right_col and bottom_row are therefore inserted as duplicates of left_col and top_row
#         cell_datatype is left empty
#     ii. calculation for left_col only works with up to 26 columns (cols A-Z)
#         will need additional logic for tables containing cols AA, AB etc

insert_stmt="INSERT INTO table_cell \
            (table_id, left_col, top_row, right_col, bottom_row, cell_content, cell_annotation) \
            SELECT  "+str(table_id)+", \
                    ascii(entry_provenance_col)-ascii('"+tablestart_col+"'), \
                    entry_provenance_row-"+str(tablestart_row)+", \
                    ascii(entry_provenance_col)-ascii('"+tablestart_col+"'), \
                    entry_provenance_row-"+str(tablestart_row)+", \
                    entry_value, 'DATA' \
            FROM entry_temp WHERE table_id="+str(table_id)
cur.execute(insert_stmt)

# Populate entry based on table_cell

insert_stmt="INSERT INTO entry (entry_cell_id) \
             SELECT cell_id FROM table_cell \
             WHERE table_id="+str(table_id)+" AND cell_annotation='DATA'"
cur.execute(insert_stmt)

# Process each row from the LABELS sheet (all_labels_rows)
# and insert the data into the temporary table label_temp
# NOTES: 
#    header row not processed - currently checking each row to see if it's the header - improve this!
#    labels containing single quotes cause errors. Either surround text in double quotes or escape the single quotes

for row in all_labels_rows:
    label_val  = str(row[0].value).replace("'","''")           # value in LABEL column with single quotes escaped
    label_val = label_val
    label_prov_text = str(row[1].value)                        # value in PROVENANCE column
    label_par  = str(row[2].value)                             # value in PARENT column
    label_cat  = str(row[3].value)                             # value in CATEGORY column
    if label_prov_text != 'PROVENANCE':                        # ignore header row
        label_prov=label_prov_text.split('","')[1].split('")')[0]      # get display val between "," and ")
        label_prov_col=''.join(filter(str.isalpha, label_prov))        # alpha part
        label_prov_row=int(''.join(filter(str.isdigit, label_prov)))   # numeric part        
        # If the label has no parent labels, insert a row containing information for just this label
        if label_par=='None':
            insert_stmt="INSERT INTO label_temp (table_id, label_value, label_provenance, label_provenance_col, label_provenance_row, label_category) \
                        VALUES ("+str(table_id)+",'"+label_val+"', '"+label_prov+"', '"+label_prov_col+"', '"+str(label_prov_row)+"', '"+label_cat+"')"
        # If the label has a parent, insert a row that includes provenance information for the parent label
        else:
            label_par_prov=label_par.split('[')[1].split(']')[0]           # get val between '[' and ']'
            insert_stmt="INSERT INTO label_temp (table_id, label_value, label_provenance, label_provenance_col, label_provenance_row, label_parent, label_category) \
                        VALUES ("+str(table_id)+",'"+label_val+"', '"+label_prov+"', '"+label_prov_col+"', '"+str(label_prov_row)+"', '"+label_par_prov+"', '"+label_cat+"')"
        cur.execute(insert_stmt)

# Populate the category table from the label_categories in label_temp

insert_stmt_cat="INSERT INTO category (category_name, table_id) \
                 SELECT DISTINCT label_category, "+str(table_id)+" FROM label_temp WHERE table_id="+str(table_id)
cur.execute(insert_stmt_cat)

# Populate table_cell based on label_temp contents
# The table_cell coordinates are calculated as for the table_cell rows corresponding to entries

insert_stmt="INSERT INTO table_cell \
            (table_id, left_col, top_row, right_col, bottom_row, cell_content, cell_annotation) \
            SELECT  "+str(table_id)+", \
                    ascii(label_provenance_col)-ascii('"+tablestart_col+"'), \
                    label_provenance_row-"+str(tablestart_row)+", \
                    ascii(label_provenance_col)-ascii('"+tablestart_col+"'), \
                    label_provenance_row-"+str(tablestart_row)+", \
                    label_value, 'HEADING' \
            FROM label_temp WHERE table_id="+str(table_id)
cur.execute(insert_stmt)

# Populate label based on table_cell and label_temp

insert_stmt="INSERT INTO label (label_cell_id, category_name, parent_label_cell_id) \
             SELECT tcv.cell_id, lt.label_category, tpcv.cell_id \
             FROM label_temp lt \
             JOIN tabby_cell_view tcv \
             ON lt.label_provenance=tcv.cell_provenance  \
             AND lt.table_id=tcv.table_id \
             LEFT JOIN tabby_cell_view tpcv \
             ON lt.label_parent=tpcv.cell_provenance  \
             AND lt.table_id = tpcv.table_id \
             WHERE lt.table_id="+str(table_id)+" AND tcv.cell_annotation='HEADING'"
cur.execute(insert_stmt)

# Populate entry_label from entry_label_temp and tabby_cell_view

# NEED TO ADD IN table_ID?

insert_el="INSERT INTO entry_label (entry_cell_id, label_cell_id) \
           SELECT e_cell.cell_id, l_cell.cell_id \
           FROM entry_label_temp elt \
           JOIN tabby_cell_view e_cell ON elt.entry_provenance = e_cell.cell_provenance \
           JOIN tabby_cell_view l_cell ON elt.label_provenance = l_cell.cell_provenance \
           WHERE elt.table_id="+str(table_id)+" \
           AND e_cell.table_id="+str(table_id)+" \
           AND l_cell.table_id="+str(table_id)
cur.execute(insert_el)

# Generate canonical_table_<table_id> and display contents

create_ctv="CALL create_tabby_canonical_table("+str(table_id)+")"
cur.execute(create_ctv)

select_ctv="SELECT * FROM tabby_canonical_table_"+str(table_id)
cur.execute(select_ctv)
canonical_table = cur.fetchall()
print(canonical_table)

# Commit changes to retain data input into tables
cur.execute('COMMIT;')
cur.close()